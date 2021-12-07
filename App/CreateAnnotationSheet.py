import xlsxwriter # cannot modify excel files
from openpyxl import load_workbook # can modify excel files
import re

def create_sheet(label_me, creation_time):

    workbook = xlsxwriter.Workbook('AL/data/annotations/' + creation_time + '.xlsx')
    worksheet = workbook.add_worksheet()

    for i in range(0, len(label_me["title"])):
        # row, column, item
        worksheet.write(i, 0, i)
        worksheet.write(i, 1, label_me.loc[[i], ["title"]].values[0][0])
        if label_me["missing_content"][i] == True:
            worksheet.write(int(i/10), 2, " ")
        else:
            worksheet.write(int(i/10), 2, str(re.sub(r'\n',' ', label_me["selftext"][i][:100])))
        worksheet.write(i, 3, "?")

    workbook.close()

    wb = load_workbook(filename = "AL/data/annotations/" + creation_time + ".xlsx")
    ws = wb.active

    ws["E2"] = "Number of annotations:"
    ws["E3"] = 0

    wb.save(filename = "AL/data/annotations/" + creation_time + ".xlsx")
    wb.close()